import os
import traceback
import pandas as pd
import pyautogui as pg


import time
# hwp
import win32com.client as win32
import win32gui
# excel
import openpyxl as op
# pdf
import shutil
# doc,ppt
import comtypes.client
# text
import aspose.words as aw

# image
from PIL import Image


class ConvertPdf:
    from_path = None # 읽을 폴더 경로
    to_path = None # 쓸 폴더 경로
    all_files = None # 모든 파일들의 이름을 담은 리스트
    def __init__(self,from_path, to_path, d):
        self.from_path = from_path + "/" + d
        self.to_path = to_path + "/" + d

        self.all_files = os.listdir(self.from_path)
        self.error_dict = {"error_file":[],"error_message":[]}

        self.run_time = 2 # 파일당 변환 시도 횟수

        if not os.path.exists(self.to_path):
            os.mkdir(self.to_path)

        for file in self.all_files:
            print(file)
            for t in range(self.run_time):
                if self.change_file_name_pdf(file) in os.listdir(self.to_path):
                    print(file," : converted already")
                    break

                if file=="PDF" or file[:2]=="~$":
                    break

                if os.path.getsize(os.path.join(self.from_path,file))==0:
                    self.error_dict["error_file"].append(file)
                    self.error_dict["error_message"].append("zero byte")
                    break

                file_type = os.path.splitext(file)[1][1:].lower()
                if file_type in ("pdf"):
                    self.pdf2pdf(file,t)
                elif file_type in ('hwp','hwpx'):
                    self.hwp2pdf_print(file,t)
                elif file_type in ('png','jpg','jpeg','jfif',"bmp"):
                    self.img2pdf(file,t)
                elif file_type in ('xlsx','xls','xlsm'):
                    self.exl2pdf_v2(file,t)
                elif file_type in ('txt'):
                    self.text2pdf(file,t)
                elif file_type in ('docx'):
                    self.word2pdf(file,t)
                elif file_type in ('ppt','pptx'):
                    self.ppt2pdf(file,t)
                else:
                    self.error_dict["error_file"].append(file)
                    self.error_dict["error_message"].append("no type converter")
                    break

    def get_file_name(self,full_name):
        return full_name.split(".")[0]

    def change_file_name_pdf(self,file_name,to_type="pdf"):
        L = file_name.split(".")
        L[-1]=to_type
        return ".".join(L)

    def to_csv_error_file(self,path):
        all_files_before = set(map(self.get_file_name,self.all_files))
        all_files_after = set(map(self.get_file_name,os.listdir(self.to_path)))
        error_files = set(map(self.get_file_name, self.error_dict["error_file"]))

        remain_error_files = list(all_files_before - all_files_after - error_files)

        if remain_error_files:
            self.error_dict["error_file"].extend(remain_error_files)
            self.error_dict["error_message"].extend([None for _ in range(len(remain_error_files))])

        pd.DataFrame(self.error_dict).to_csv(path,encoding="utf-8",index=False)

    def text2pdf(self,file_name,t):
        try:
            doc = aw.Document(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            doc.save(os.path.join(self.to_path, pdf_file_name))
        except:
            if t == self.all_files - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def word2pdf(self,file_name,t):
        try:
            word = comtypes.client.CreateObject('Word.Application')
            word.Visible = True

            doc = word.Documents.Open(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            doc.SaveAs(os.path.join(self.to_path, pdf_file_name), FileFormat=17)
            doc.Close()
            word.Quit()
        except:
            try:
                doc.Close()
            except:
                pass
            try:
                word.Quit()
            except:
                pass

            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def ppt2pdf(self,file_name,t):
        try:
            powerpoint = comtypes.client.CreateObject("Powerpoint.Application")
            powerpoint.Visible = True

            slides = powerpoint.Presentations.Open(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            full_to_path = os.path.join(self.to_path, pdf_file_name).replace("/", "\\")

            slides.SaveAs(full_to_path, FileFormat=32)
            slides.Close()
            powerpoint.Quit()
        except:
            try:
                slides.Close()
            except:
                pass
            try:
                powerpoint.Quit()
            except:
                pass

            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def pdf2pdf(self,file_name,t):
        try:
            pdf_file_name = self.change_file_name_pdf(file_name)
            shutil.copyfile(os.path.join(self.from_path, file_name),os.path.join(self.to_path, pdf_file_name))
            # print(file_name, "=>", file_name)
        except:
            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def hwp2pdf(self, file_name, t):
        try:
            hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")  # hwp 창열기
            hwp.RegisterModule('FilePathCheckDLL', 'AutomationModule')  # 보안모듈 삭제
            win32gui.FindWindow(None, 'Noname 1 - HWP')

            hwp.Open(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            hwp.SaveAs(os.path.join(self.to_path, pdf_file_name), "PDF")
            # print(file_name, "=>", pdf_file_name)
            hwp.Quit()
        except:
            try:
                hwp.Quit()
            except:
                pass

            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def hwp2pdf_print(self, file_name, t):


        try:
            hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")  # hwp 창열기
            hwp.RegisterModule('FilePathCheckDLL', 'AutomationModule')
            win32gui.FindWindow(None, 'Noname 1 - HWP')

            hwp.Open(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)

            act = hwp.CreateAction("Print")
            pset = act.CreateSet()
            act.GetDefault(pset)
            pset.SetItem("PrintMethod",0) # 0:보통출력, 4:2쪽모아찍기
            pset.SetItem("FileName",os.path.join(self.to_path, pdf_file_name))
            pset.SetItem("printerName","Hancom PDF")
            act.Execute(pset)

            hwp.Quit()
            # hwp.XHwpDocuments.Item(0).XHwpPrint.filename = os.path.join(self.to_path, pdf_file_name)
            # hwp.XHwpDocuments.Item(0).XHwpPrint.RunToPDF()
            # hwp.Quit()
        except:
            try:
                hwp.Quit()
            except:
                pass

            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def img2pdf(self,file_name,t):
        try:
            im = Image.open(os.path.join(self.from_path,file_name)).convert("RGB")
            pdf_file_name = self.change_file_name_pdf(file_name)
            im.save(os.path.join(self.to_path,pdf_file_name),save_all=True)
            # print(file_name, "=>", pdf_file_name)
        except:
            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    # def exl2pdf_print(self,file_name,t):
    #     try:
    #         if file_name.endswith("xls"):
    #             excel2 = win32.gencache.EnsureDispatch('Excel.application')
    #             wb = excel2.Workbooks.Open(os.path.join(self.from_path, file_name))
    #             xlsx_file_name = self.change_file_name_pdf(file_name, "xlsx")
    #
    #             full_from_path = os.path.join(self.from_path, xlsx_file_name).replace("/", "\\")
    #             wb.SaveAs(full_from_path, FileFormat=51)
    #             wb.Close()
    #             excel2.Application.Quit()
    #
    #             os.remove(os.path.join(self.from_path, file_name))
    #             file_name = xlsx_file_name
    #
    #         wb = op.load_workbook(self.from_path + "/" + file_name)  # openptxl workbook생성
    #
    #         # 활성시트만 저장
    #         ws_list = []
    #         xls = pd.ExcelFile(self.from_path + "/" + file_name)
    #         sheets = xls.book.sheets()
    #         for sheet in sheets:
    #             if sheet.visibility==0:
    #                 ws_list.append(sheet.name)
    #
    #         excel = win32.Dispatch("Excel.Application")
    #         wb = excel.Workbooks.Open(self.from_path + "/" + file_name)
    #
    #         for ws in ws_list:
    #             wb.Worksheets(ws_list).Select()
    #
    #         xlsx_file_name = self.change_file_name_pdf(file_name, "pdf")
    #         wb.ActiveSheet.ExportAsFixedFormat(0, self.to_path + "/" + xlsx_file_name)  # 파일명, 시트명으로 pdf 파일 저장
    #
    #         wb.Close(False)  # workbook 닫기. True일 경우 그 상태를 저장한다.
    #         excel.Quit()  # excel 닫기
    #     except:
    #         try:
    #             wb.Close(False)
    #         except:
    #             pass
    #         try:
    #             excel.Quit()
    #         except:
    #             pass
    #
    #         if t == self.run_time - 1:
    #             self.error_dict["error_file"].append(file_name)
    #             self.error_dict["error_message"].append(traceback.format_exc())

    def exl2pdf_v2(self,file_name,t):
        try:
            if file_name.endswith("xls"):
                excel2 = win32.gencache.EnsureDispatch('Excel.application')
                wb = excel2.Workbooks.Open(os.path.join(self.from_path, file_name))
                xlsx_file_name = self.change_file_name_pdf(file_name, "xlsx")

                full_from_path = os.path.join(self.from_path, xlsx_file_name).replace("/", "\\")
                wb.SaveAs(full_from_path, FileFormat=51)
                wb.Close()
                excel2.Application.Quit()

                os.remove(os.path.join(self.from_path, file_name))
                file_name = xlsx_file_name

            # 활성시트만 저장
            def get_active_sheets():
                ws_list = []
                wb = op.load_workbook(self.from_path + "/" + file_name)
                for sheet in wb.worksheets:
                    if sheet.sheet_state=="visible":
                        ws_list.append(sheet.title)
                return ws_list

            excel = win32.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(self.from_path + "/" + file_name)
            ws_list = get_active_sheets()
            wb.Worksheets(ws_list).Select()

            xlsx_file_name = self.change_file_name_pdf(file_name, "pdf")
            wb.ActiveSheet.ExportAsFixedFormat(0, self.to_path + "/" + xlsx_file_name)  # 파일명, 시트명으로 pdf 파일 저장

            wb.Close(False)  # workbook 닫기. True일 경우 그 상태를 저장한다.
            excel.Quit()  # excel 닫기
        except:
            try:
                wb.Close(False)
            except:
                pass
            try:
                excel.Quit()
            except:
                pass

            if t == self.run_time - 1:
                self.error_dict["error_file"].append(file_name)
                self.error_dict["error_message"].append(traceback.format_exc())

    def exl2pdf(self):
        def excelInfo(filepath):
            excel_list = [file for file in os.listdir(filepath) if
                          file.endswith('xlsx') and file[:2] != "~$"]  # 폴더안에있는 엑셀파일 명을 리스트로 저장
            result = []  # 빈 리스트 생성

            for file in excel_list:  # 엑셀파일명 리스트를 for문을 통해 반복
                wb = op.load_workbook(filepath + "/" + file)  # openptxl workbook생성
                ws_list = wb.sheetnames  # 해당 workbook의 시트명을 리스트로 받음
                filename = file.replace(".xlsx", "")  # 파일명을 저장하기 위해 문자열에서 확장자를 제거

                for sht in ws_list:  # 시트명 리스트를 for문을 통해 반복
                    temp_tuple = (filepath + "/" + file, filename, sht)  # 파일경로, 파일명, sht를 튜플에 저장
                    result.append(temp_tuple)  # 위 튜플을 빈 리스트에 추가

            return result  # 튜플로 이루어진 리스트 리턴

        def transPDF(fileinfo, savepath):
            excel = win32.Dispatch("Excel.Application")
            i = 0  # 파일명 중복을 방지하기 위한 인덱싱 번호
            # excelinfo를 받아서 for문을 실행
            for info in fileinfo:
                wb = excel.Workbooks.Open(info[0])  # info가 튜플이므로 인덱싱으로 접근(0번째는 파일경로)
                ws = wb.Worksheets(info[2])  # 튜플의 2번쨰 요소는 시트명
                ws.Select()  # 위 설정한 시트 선택
                try:
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath + "/" + str(i) + "_" + info[1] + "_" + info[2] + ".pdf")  # 파일명, 시트명으로 pdf 파일 저장
                except:
                    pass
                i = i + 1
                wb.Close(False)  # workbook 닫기. True일 경우 그 상태를 저장한다.
                excel.Quit()  # excel 닫기

        excelinfo = excelInfo(self.from_path)
        transPDF(excelinfo, self.to_path)